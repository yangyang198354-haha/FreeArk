import os
import sys
import json
import time
from typing import Dict, List, Any
import concurrent.futures
import pandas as pd
import copy

# 处理PyInstaller打包后的资源文件路径
def get_resource_path(relative_path):
    """获取资源文件的绝对路径，支持PyInstaller打包环境"""
    try:
        # PyInstaller打包后的临时目录
        base_path = sys._MEIPASS
    except Exception:
        # 正常开发环境
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    return os.path.join(base_path, relative_path)

# 添加FreeArk目录到Python路径，确保模块可以正确导入
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# 导入统一的日志配置管理器
from datacollection.log_config_manager import get_logger

# 导入PLC读取相关类
from datacollection.multi_thread_plc_handler import PLCReadWriter, PLCManager
# 导入MQTT客户端
from datacollection.mqtt_client import MQTTClient

# 获取logger，日志级别从配置文件读取
logger = get_logger('improved_data_collection')

class ImprovedDataCollectionManager:
    def _get_resource_dir(self):
        """获取资源目录，支持多种运行环境"""
        # 尝试从多个位置获取资源目录
        possible_dirs = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resource'),  # 模块同级resource目录（含完整参数）
            os.path.join(os.getcwd(), 'resource'),  # 当前工作目录下的resource
            os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'resource'),  # 项目根resource目录
        ]
        
        # 优先选择存在的目录
        for dir_path in possible_dirs:
            if os.path.exists(dir_path) and os.path.isdir(dir_path):
                return dir_path
        
        # 如果都不存在，返回当前工作目录
        return os.getcwd()
    
    def _get_output_dir(self):
        """获取输出目录"""
        possible_dirs = [
            os.path.join(os.getcwd(), 'output'),  # 当前工作目录下的output
            os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'output'),  # 项目output目录
        ]
        
        # 优先选择存在的目录
        for dir_path in possible_dirs:
            try:
                if not os.path.exists(dir_path):
                    os.makedirs(dir_path)
                return dir_path
            except:
                continue
        
        # 如果都不行，使用临时目录
        import tempfile
        return tempfile.gettempdir()
        
    def __init__(self, max_workers: int = 10):
        """初始化改进的数据收集管理器"""
        self.max_workers = max_workers
        self.plc_manager = PLCManager(max_workers=max_workers)
        # 使用辅助方法获取目录
        self.resource_dir = self._get_resource_dir()
        self.output_dir = self._get_output_dir()
        # 确保output目录存在
        if not os.path.exists(self.output_dir):
            try:
                os.makedirs(self.output_dir)
            except:
                pass
        self.results = {}

    def start(self):
        """启动数据收集管理器"""
        self.plc_manager.start()
        logger.info(f"✅ 改进版数据收集管理器已启动，线程池大小：{self.max_workers}")

    def stop(self):
        """停止数据收集管理器"""
        self.plc_manager.stop()
        logger.info("✅ 改进版数据收集管理器已停止")
    
    def load_building_json(self, building_file: str) -> Dict[str, Dict[str, Any]]:
        """加载楼栋的JSON文件"""
        # 尝试多种路径
        possible_paths = [
            os.path.join(self.resource_dir, building_file),  # 资源目录
            get_resource_path(building_file),  # 使用通用路径函数
            get_resource_path(os.path.join('resource', building_file)),  # resource子目录
            building_file  # 直接使用传入的路径
        ]
        
        # 尝试从可能的路径加载文件
        for file_path in possible_paths:
            if os.path.exists(file_path):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        logger.info(f"✅ 成功加载楼栋JSON文件：{building_file}，共{len(data)}条记录")
                        return data
                except Exception as e:
                    logger.info(f"❌ 从{file_path}加载楼栋JSON文件失败：{str(e)}")
                    continue
        
        logger.info(f"❌ 未找到楼栋JSON文件：{building_file}")
        return {}

    def load_plc_config(self) -> Dict[str, Dict[str, Any]]:
        """加载PLC配置文件"""
        # 尝试多种路径
        possible_paths = [
            os.path.join(self.resource_dir, 'plc_config.json'),  # 资源目录
            get_resource_path('plc_config.json'),  # 使用通用路径函数
            get_resource_path(os.path.join('resource', 'plc_config.json')),  # resource子目录
            os.path.join(os.getcwd(), 'plc_config.json')  # 当前工作目录
        ]
        
        # 尝试从可能的路径加载文件
        for config_path in possible_paths:
            if os.path.exists(config_path):
                try:
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                        logger.info(f"✅ 成功加载PLC配置文件，包含{len(config.get('parameters', {}))}个参数")
                        return config.get('parameters', {})
                except Exception as e:
                    logger.info(f"❌ 从{config_path}加载PLC配置文件失败：{str(e)}")
                    continue
        
        logger.info(f"❌ 未找到PLC配置文件")
        # 返回默认配置
        return {}
    
    def load_output_config(self) -> Dict[str, Any]:
        """加载输出配置文件"""
        # 尝试多种路径
        possible_paths = [
            os.path.join(self.resource_dir, 'output_config.json'),  # 资源目录
            get_resource_path('output_config.json'),  # 使用通用路径函数
            get_resource_path(os.path.join('resource', 'output_config.json')),  # resource子目录
            os.path.join(os.getcwd(), 'output_config.json')  # 当前工作目录
        ]
        
        # 默认配置
        default_config = {
            "output": {
                "type": "Excel",
                "excel": {
                    "file_name": "累计用量",
                    "directory": self.output_dir,
                    "include_all_params": True
                },
                "json": {
                    "enabled": True
                },
                "mqtt": {
                    "enabled": False,
                    "server": {
                        "host": "localhost",
                        "port": 1883
                    }
                }
            }
        }
        
        # 尝试从可能的路径加载文件
        for config_path in possible_paths:
            if os.path.exists(config_path):
                try:
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                        logger.info(f"✅ 成功加载输出配置文件")
                        # 处理相对路径，确保指向正确的输出目录
                        if 'output' in config and 'excel' in config['output'] and 'directory' in config['output']['excel']:
                            directory = config['output']['excel']['directory']
                            # 如果是相对路径，转换为绝对路径
                            if not os.path.isabs(directory):
                                # 使用项目的output目录作为基准
                                config['output']['excel']['directory'] = os.path.join(self.output_dir, directory.lstrip('./\\'))
                        return config
                except Exception as e:
                    logger.info(f"❌ 从{config_path}加载输出配置文件失败：{str(e)}")
                    continue
        
        logger.info(f"❌ 未找到输出配置文件，使用默认配置")
        return default_config

    # 移除了load_room_plc_map方法，因为房间与PLC IP映射文件已不存在
    
    def load_output_config(self) -> Dict[str, Any]:
        """加载输出配置文件"""
        config_path = os.path.join(self.resource_dir, 'output_config.json')
        # 默认配置
        default_config = {
            "output": {
                "type": "Excel",
                "excel": {
                    "file_name": "累计用量",
                    "directory": self.output_dir,
                    "include_all_params": True
                },
                "json": {
                    "enabled": True
                },
                "mqtt": {
                    "enabled": False,
                    "server": {
                        "host": "localhost",
                        "port": 1883,
                        "username": "",
                        "password": "",
                        "tls_enabled": False
                    },
                    "topic": {
                        "prefix": "/datacollection/plc/to/collector/"
                    },
                    "qos": 1,
                    "retain": False
                }
            }
        }
        
        if not os.path.exists(config_path):
            logger.info(f"⚠️  输出配置文件不存在，使用默认配置：{config_path}")
            return default_config
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                logger.info(f"✅ 成功加载输出配置文件")
                # 合并默认配置和文件配置
                if 'output' not in config:
                    config['output'] = default_config['output']
                else:
                    if 'excel' not in config['output']:
                        config['output']['excel'] = default_config['output']['excel']
                    else:
                        config['output']['excel'] = {**default_config['output']['excel'], **config['output']['excel']}
                    if 'json' not in config['output']:
                        config['output']['json'] = default_config['output']['json']
                    else:
                        config['output']['json'] = {**default_config['output']['json'], **config['output']['json']}
                    if 'mqtt' not in config['output']:
                        config['output']['mqtt'] = default_config['output']['mqtt']
                    else:
                        # 合并MQTT配置，确保所有必需的子项都存在
                        if 'server' not in config['output']['mqtt']:
                            config['output']['mqtt']['server'] = default_config['output']['mqtt']['server']
                        else:
                            config['output']['mqtt']['server'] = {**default_config['output']['mqtt']['server'], **config['output']['mqtt']['server']}
                        if 'topic' not in config['output']['mqtt']:
                            config['output']['mqtt']['topic'] = default_config['output']['mqtt']['topic']
                        else:
                            config['output']['mqtt']['topic'] = {**default_config['output']['mqtt']['topic'], **config['output']['mqtt']['topic']}
                        # 合并其他MQTT配置项
                        config['output']['mqtt'] = {**default_config['output']['mqtt'], **config['output']['mqtt']}
                return config
        except Exception as e:
            logger.info(f"❌ 加载输出配置文件失败，使用默认配置：{str(e)}")
            return default_config

    def collect_data_for_building(self, building_file: str,
                                    param_filter: set = None) -> Dict[str, Dict[str, Any]]:
        """为指定楼栋收集数据，使用PLC IP地址而不是设备IP地址。

        Args:
            building_file: 楼栋JSON文件名（相对于resource目录）
            param_filter:  若提供，则只采集该集合内的参数名；为 None 时采集全部参数。
        """
        # 加载楼栋数据和PLC配置
        building_data = self.load_building_json(building_file)
        if not building_data:
            return {}

        plc_config = self.load_plc_config()
        if not plc_config:
            return {}

        # 若指定了参数过滤集合，只保留其中存在于 plc_config 的参数
        if param_filter is not None:
            plc_config = {k: v for k, v in plc_config.items() if k in param_filter}
            if not plc_config:
                logger.warning(f"⚠️  param_filter 过滤后无有效参数，跳过楼栋 {building_file}")
                return {}

        # 创建PLC读取配置列表
        plc_read_configs = []
        ip_to_device_map = {}

        # 为每个设备的每个参数创建读取配置
        for device_id, device_info in building_data.items():

            # 优先使用设备信息中的PLC IP地址
            plc_ip = device_info.get('PLC IP地址')

            if not plc_ip:
                # 如果仍然没有PLC IP，使用设备的IP地址作为后备方案
                plc_ip = device_info.get('IP地址')
                if not plc_ip:
                    logger.info(f"⚠️  设备 {device_id} 没有可用的IP地址")
                    continue
                logger.info(f"⚠️  设备 {device_id} 没有PLC IP，使用设备IP: {plc_ip}")

            # 为每个参数创建配置
            for param_key, param_info in plc_config.items():
                config = {
                    'ip': plc_ip,
                    'db_num': param_info.get('db_num'),
                    'offset': param_info.get('offset'),
                    'length': param_info.get('length'),
                    'data_type': param_info.get('data_type'),
                    'device_id': device_id,
                    'param_key': param_key,
                    'original_device_ip': device_info.get('IP地址')
                }
                plc_read_configs.append(config)

                # 记录PLC IP到设备的映射
                if plc_ip not in ip_to_device_map:
                    ip_to_device_map[plc_ip] = []
                ip_to_device_map[plc_ip].append(device_id)
        
        logger.info(f"🚀 开始为楼栋 {building_file} 收集数据，共{len(plc_read_configs)}个读取任务，涉及{len(ip_to_device_map)}个PLC设备...")
        start_time = time.time()
        
        # 读取所有PLC数据
        results = self._read_all_plc_data(plc_read_configs)
        
        # 组织结果
        organized_results = self._organize_results(results, building_data, plc_config)
        
        elapsed_time = time.time() - start_time
        logger.info(f"⏱️  数据收集完成，耗时：{elapsed_time:.2f} 秒")
        
        # 保存结果
        self.results[building_file] = organized_results
        
        # 获取输出配置
        output_config = self.load_output_config()
        output_type = output_config['output'].get('type', 'Excel')
        
        # 获取各种输出方式的enabled配置
        json_config = output_config['output'].get('json', {})
        json_enabled = json_config.get('enabled', True)
        
        excel_config = output_config['output'].get('excel', {})
        excel_enabled = excel_config.get('enabled', True)
        
        mqtt_config = output_config['output'].get('mqtt', {})
        mqtt_enabled = mqtt_config.get('enabled', False)
        
        # 根据配置保存结果
        # 如果输出类型为Json或者JSON输出已启用，则保存为JSON文件
        if output_type == 'Json' or json_enabled:
            # 保存为JSON文件
            self.save_results_to_json(building_file)
        
        # 如果输出类型为Excel或者Excel输出已启用，则保存为Excel文件
        if output_type == 'Excel' or excel_enabled:
            # 保存为Excel文件
            self.save_results_to_excel(building_file)
        
        # 如果输出类型为MQTT或者MQTT输出已启用，则通过MQTT发送数据
        if output_type == 'MQTT' or mqtt_enabled:
            # 直接传入 organized_results，避免 self.results 被其他线程覆盖导致发送错误数据
            self.send_results_to_mqtt(building_file, results_data=organized_results)
        
        return organized_results

    def _read_all_plc_data(self, plc_read_configs: List[Dict]) -> List[Dict]:
        """读取所有PLC数据，按IP分组后并行提交线程池，各IP内部分块批量读取"""
        # 按PLC IP地址对参数配置进行分组
        ip_to_configs = {}
        for config in plc_read_configs:
            plc_ip = config['ip']
            if plc_ip not in ip_to_configs:
                ip_to_configs[plc_ip] = []
            ip_to_configs[plc_ip].append(config)

        # 通过 PLCManager 线程池提交任务（每个IP一个任务，内部分块+连接复用）
        future_to_ip = {}
        for plc_ip, configs in ip_to_configs.items():
            future = self.plc_manager.thread_pool.submit(
                self._read_single_plc_with_multiple_params, plc_ip, configs
            )
            future_to_ip[future] = plc_ip

        # 收集结果
        results = []
        for future in concurrent.futures.as_completed(future_to_ip):
            plc_ip = future_to_ip[future]
            try:
                ip_results = future.result(timeout=60)
                results.extend(ip_results)
            except concurrent.futures.TimeoutError:
                logger.info(f"❌ PLC任务超时（60s）：{plc_ip}")
                for config in ip_to_configs.get(plc_ip, []):
                    results.append({
                        'ip': config['ip'],
                        'device_id': config.get('device_id'),
                        'param_key': config.get('param_key'),
                        'success': False,
                        'message': "任务执行超时（60秒）",
                        'value': None
                    })
            except Exception as e:
                logger.info(f"❌ PLC任务执行异常：{plc_ip} - {str(e)}")
                for config in ip_to_configs.get(plc_ip, []):
                    results.append({
                        'ip': config['ip'],
                        'device_id': config.get('device_id'),
                        'param_key': config.get('param_key'),
                        'success': False,
                        'message': f"任务执行异常：{str(e)}",
                        'value': None
                    })

        return results

    def _read_single_plc_with_multiple_params(self, plc_ip: str, configs: List[Dict]) -> List[Dict]:
        """读取单个PLC的多个参数——委托给 PLCManager 以复用连接并进行分块批量读取。

        PLCManager._read_single_plc_multiple_params 已实现：
          - clients_cache 连接复用（同一 IP 跨轮次不断开）
          - PDU_CHUNK_SIZE 分块（≤12 参数/请求）
        本方法额外将 PLCManager 结果格式转换为包含 device_id / param_key 的格式。
        """
        # PLCManager 方法返回的结果键为 db_num/offset，不含 device_id/param_key
        # 先建立 (db_num, offset) -> (device_id, param_key) 映射
        key_map = {}
        for config in configs:
            k = (config['db_num'], config['offset'])
            key_map[k] = (config.get('device_id'), config.get('param_key'))

        # 统计参与本次读取的设备数和参数数
        device_ids_in_configs = set(c.get('device_id') for c in configs if c.get('device_id'))
        logger.debug(
            f"[_read_single_plc] plc_ip={plc_ip} key_map 构建完成: "
            f"设备数={len(device_ids_in_configs)}, 参数数={len(key_map)}"
        )

        raw_results = self.plc_manager._read_single_plc_multiple_params(plc_ip, configs)

        results = []
        for i, raw in enumerate(raw_results):
            config = configs[i]
            db_num = config['db_num']
            offset = config['offset']
            device_id, param_key = key_map.get((db_num, offset), (None, None))
            if device_id is None or param_key is None:
                logger.warning(
                    f"[_read_single_plc] key_map 未命中: plc_ip={plc_ip}, "
                    f"db_num={db_num}, offset={offset} -> device_id={device_id}, param_key={param_key}"
                )
            results.append({
                'ip': plc_ip,
                'device_id': device_id,
                'param_key': param_key,
                'success': raw.get('success', False),
                'message': raw.get('message', ''),
                'value': raw.get('value')
            })
        return results
    
    # 原有的_read_single_plc_with_param方法可以保留或删除，根据是否有其他地方调用决定
    def _read_single_plc_with_param(self, config: Dict) -> Dict:
        """读取单个PLC的单个参数"""
        plc_ip = config['ip']
        db_num = config['db_num']
        offset = config['offset']
        length = config['length']
        data_type = config['data_type']
        device_id = config.get('device_id')
        param_key = config.get('param_key')
        
        # 导入PLC读取器类
        from datacollection.multi_thread_plc_handler import PLCReadWriter
        # 创建PLC读取器并连接
        reader = PLCReadWriter(plc_ip)
        try:
            if not reader.connect():
                # 只连接PLC IP，如果失败直接标记为失败
                logger.info(f"❌ PLC IP连接失败: {plc_ip}")
                return {
                    'ip': plc_ip,
                    'device_id': device_id,
                    'param_key': param_key,
                    'success': False,
                    'message': "PLC IP连接失败",
                    'value': None
                }
            
            # 读取数据
            success, message, value = reader.read_db_data(db_num, offset, length, data_type)
            return {
                'ip': plc_ip,
                'device_id': device_id,
                'param_key': param_key,
                'success': success,
                'message': message,
                'value': value
            }
        finally:
            # 确保断开连接
            reader.disconnect()

    def _organize_results(self, results: List[Dict], building_data: Dict, plc_config: Dict) -> Dict[str, Dict[str, Any]]:
        """组织结果数据"""
        organized_results = {} 
        success_count = 0
        total_count = len(results)
        
        # 获取当前格式化的时间字符串，用于所有参数
        current_time_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        
        # 初始化所有设备的结果
        for device_id, device_info in building_data.items():
            organized_results[device_id] = {
                **device_info,  # 复制原始设备信息
                'data': {},  # 添加数据字段
                'status': 'pending',  # 初始状态
                'timestamp': current_time_str  # 为设备添加时间戳
            }
        
        # 处理每个结果
        for result in results:
            device_id = result.get('device_id')
            param_key = result.get('param_key')

            if device_id and device_id in organized_results and param_key:
                # 存储参数结果，添加时间戳
                organized_results[device_id]['data'][param_key] = {
                    'value': result.get('value'),
                    'success': result.get('success'),
                    'message': result.get('message'),
                    'timestamp': current_time_str  # 为每个参数添加时间戳
                }

                # 更新设备状态
                if result.get('success'):
                    organized_results[device_id]['status'] = 'success'
                    success_count += 1
                    # 打印成功读取的日志
                    logger.info(f"✅ 设备 {device_id} 参数 {param_key} 读取成功，值：{result.get('value')}")
                else:
                    organized_results[device_id]['status'] = 'partial_success' if organized_results[device_id]['status'] == 'success' else 'failed'
                    logger.debug(
                        f"[_organize_results] 参数读取失败: device_id={device_id}, "
                        f"param_key={param_key}, message={result.get('message', '')}"
                    )

        logger.info(f"📊 数据收集结果统计：成功 {success_count}/{total_count} 个参数读取任务")

        # 按设备汇总成功/失败参数数量
        for dev_id, dev_data in organized_results.items():
            params = dev_data.get('data', {})
            dev_success = sum(1 for p in params.values() if p.get('success'))
            dev_fail = len(params) - dev_success
            logger.debug(
                f"[_organize_results] 设备汇总: device_id={dev_id}, "
                f"成功参数数={dev_success}, 失败参数数={dev_fail}"
            )

        return organized_results

    def save_results_to_json(self, building_file: str, output_file: str = None) -> bool:
        """保存结果到JSON文件"""
        # 获取输出配置
        output_config = self.load_output_config()
        json_config = output_config['output'].get('json', {})
        json_enabled = json_config.get('enabled', True)
        
        # 检查JSON输出是否启用
        if not json_enabled:
            logger.info(f"⚠️  JSON输出未启用")
            return False
            
        if building_file not in self.results:
            logger.info(f"❌ 没有找到楼栋 {building_file} 的结果数据")
            return False
        
        # 确定输出文件名
        if not output_file:
            base_name = os.path.splitext(building_file)[0]
            # 添加时间戳信息，格式为：YYYYMMDD_HHMMSS
            timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
            output_file = f"{base_name}_improved_data_collected_{timestamp}.json"
        
        # 保存到output目录
        output_path = os.path.join(self.output_dir, output_file)
        
        try:
            # 深拷贝结果数据，避免修改原始数据
            results_copy = copy.deepcopy(self.results[building_file])
            
            # 处理时间戳：移除大整数timestamp字段，将timestamp_readable重命名为timestamp
            for device_id, device_info in results_copy.items():
                if 'data' in device_info:
                    for param_key, param_data in device_info['data'].items():
                        if 'timestamp_readable' in param_data:
                            # 保留timestamp_readable并将其重命名为timestamp
                            param_data['timestamp'] = param_data.pop('timestamp_readable', '')
                        elif 'timestamp' in param_data:
                            # 只保留timestamp字段
                            pass  # 如果没有timestamp_readable，则保持原timestamp不变
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(results_copy, f, ensure_ascii=False, indent=2)
            logger.info(f"✅ 改进版结果已保存到：{output_path}")
            return True
        except Exception as e:
            logger.info(f"❌ 保存改进版结果失败：{str(e)}")
            return False
    
    def send_results_to_mqtt(self, building_file: str, results_data: dict = None) -> bool:
        """通过MQTT发送结果数据，为每条记录单独发送消息"""
        # 获取输出配置
        output_config = self.load_output_config()
        mqtt_config = output_config['output'].get('mqtt', {})

        # 检查MQTT是否启用
        mqtt_enabled = mqtt_config.get('enabled', False)
        if not mqtt_enabled:
            logger.info(f"⚠️  MQTT输出未启用")
            return False

        if results_data is None:
            if building_file not in self.results:
                logger.info(f"❌ 没有找到楼栋 {building_file} 的结果数据")
                return False
        
        # 获取MQTT服务器配置
        server_config = mqtt_config.get('server', {})
        host = server_config.get('host', 'localhost')
        port = server_config.get('port', 1883)
        username = server_config.get('username', '')
        password = server_config.get('password', '')
        tls_enabled = server_config.get('tls_enabled', False)
        pool_size = server_config.get('pool_size', 5)
        
        # 获取MQTT主题配置
        topic_config = mqtt_config.get('topic', {})
        topic_prefix = topic_config.get('prefix', '/datacollection/plc/to/collector/')
        
        # 获取其他MQTT配置
        qos = mqtt_config.get('qos', 1)
        retain = mqtt_config.get('retain', False)
        
        try:
            # 获取结果数据：优先使用调用方传入的 results_data（避免竞争条件）
            results = results_data if results_data is not None else self.results[building_file]

            # 深拷贝结果数据，避免修改原始数据
            results_copy = copy.deepcopy(results)
            
            # 创建连接池配置
            pool_config = {
                'host': host,
                'port': port,
                'username': username if username else None,
                'password': password if password else None,
                'tls_enabled': tls_enabled,
                'pool_size': pool_size
            }
            
            # 使用MQTT客户端管理器获取连接池
            import sys
            import os
            # 添加项目根目录到Python路径
            sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            from datacollection.mqtt_client_pool import MQTTClientManager
            mqtt_manager = MQTTClientManager.get_instance(pool_config)
            
            mqtt_client = None
            total_records = len(results_copy)
            success_count = 0
            
            try:
                # 从连接池获取客户端
                mqtt_client = mqtt_manager.get_client()
                
                # 遍历每条记录，单独发送MQTT消息
                for device_id, device_info in results_copy.items():
                    # 获取当前记录的唯一标识符
                    unique_identifier = device_info.get('唯一标识符', '')
                    
                    # 如果没有找到唯一标识符，使用设备ID和时间戳作为备选
                    if not unique_identifier:
                        unique_identifier = f"{device_id}_{int(time.time())}"
                        logger.info(f"⚠️  设备 {device_id} 未找到唯一标识符，使用设备ID和时间戳代替: {unique_identifier}")
                    
                    # 构建完整的MQTT主题
                    mqtt_topic = f"{topic_prefix}{unique_identifier}"
                    
                    # 创建只包含当前设备信息的消息数据
                    single_record = {device_id: device_info}
                    
                    # 发布前记录 DEBUG 信息
                    payload_bytes = len(json.dumps(single_record, ensure_ascii=False).encode('utf-8'))
                    dev_params = device_info.get('data', {})
                    dev_success_params = sum(1 for p in dev_params.values() if isinstance(p, dict) and p.get('success'))
                    dev_fail_params = len(dev_params) - dev_success_params
                    logger.debug(
                        f"[send_results_to_mqtt] 准备发布: device_id={device_id}, "
                        f"unique_identifier={unique_identifier}, payload_size={payload_bytes}bytes, "
                        f"success_params={dev_success_params}, failed_params={dev_fail_params}"
                    )

                    # 发送数据
                    success = mqtt_client.publish(mqtt_topic, single_record, qos=qos, retain=retain)

                    if success:
                        success_count += 1
                        logger.info(f"✅ 设备 {device_id} 数据已成功发送到MQTT主题: {mqtt_topic}")
                    else:
                        logger.warning(
                            f"[send_results_to_mqtt] 发布失败: device_id={device_id}, "
                            f"topic={mqtt_topic}, payload_size={payload_bytes}bytes"
                        )
                        logger.info(f"❌ 设备 {device_id} 数据发送到MQTT主题失败: {mqtt_topic}")
                
                # 返回整体发送结果
                if success_count == total_records:
                    logger.info(f"✅ 所有 {total_records} 条记录均已成功发送到MQTT")
                    return True
                elif success_count > 0:
                    logger.info(f"⚠️  部分记录发送成功: {success_count}/{total_records}")
                    return True  # 即使部分失败，也返回True表示至少有部分成功发送
                else:
                    logger.info(f"❌ 所有 {total_records} 条记录发送失败")
                    return False
            finally:
                # 确保将客户端归还到连接池
                if mqtt_client:
                    mqtt_manager.return_client(mqtt_client)
        except Exception as e:
            logger.info(f"❌ 通过MQTT发送数据异常: {str(e)}")
            return False
    
    def save_results_to_excel(self, building_file: str) -> bool:
        """保存结果到Excel文件，将成功结果输出到success工作表，失败结果输出到failure工作表"""
        # 获取输出配置
        output_config = self.load_output_config()
        excel_config = output_config['output'].get('excel', {})
        
        # 检查Excel输出是否启用
        excel_enabled = excel_config.get('enabled', True)
        if not excel_enabled:
            logger.info(f"⚠️  Excel输出未启用")
            return False
            
        if building_file not in self.results:
            logger.info(f"❌ 没有找到楼栋 {building_file} 的结果数据")
            return False
        
        file_name = excel_config.get('file_name', '累计用量')
        directory = excel_config.get('directory', self.output_dir)
        
        # 确保输出目录存在
        os.makedirs(directory, exist_ok=True)
        
        # 生成文件名，包含时间戳
        timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
        output_file = f"{file_name}_{timestamp}.xlsx"
        output_path = os.path.join(directory, output_file)
        
        try:
            # 准备成功和失败的数据列表
            success_data = []
            failure_data = []
            results = self.results[building_file]
            
            # 获取当前格式化的时间字符串
            current_time_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
            
            # 遍历每个设备的结果，分离成功和失败的数据
            for device_id, device_info in results.items():
                # 基础信息
                device_data = device_info.copy()
                data_section = device_data.pop('data', {})
                
                # 提取每个参数的值
                for param_key, param_value in data_section.items():
                    row = {
                        'device_id': device_id,
                        'param_key': param_key,
                        'value': param_value.get('value'),
                        'success': param_value.get('success'),
                        'message': param_value.get('message'),
                        'timestamp': current_time_str  # 使用当前格式化的时间字符串，不含毫秒
                    }
                    
                    # 添加设备基本信息
                    for key, value in device_data.items():
                        row[key] = value
                    
                    # 根据成功状态分别添加到不同的数据列表
                    if param_value.get('success'):
                        success_data.append(row)
                    else:
                        failure_data.append(row)
            
            # 使用ExcelWriter来应用格式
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 导入openpyxl样式类
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                # 定义样式
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                normal_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # 保存成功数据到success工作表
                if success_data:
                    success_df = pd.DataFrame(success_data)
                    success_df.to_excel(writer, index=False, sheet_name='success')
                    success_ws = writer.sheets['success']
                    self._apply_excel_formatting(success_ws, header_font, header_fill, header_alignment, thin_border, normal_alignment)
                
                # 保存失败数据到failure工作表
                if failure_data:
                    failure_df = pd.DataFrame(failure_data)
                    failure_df.to_excel(writer, index=False, sheet_name='failure')
                    failure_ws = writer.sheets['failure']
                    self._apply_excel_formatting(failure_ws, header_font, header_fill, header_alignment, thin_border, normal_alignment)
                
                # 如果没有成功数据，至少创建一个工作表避免Excel文件为空
                if not success_data and not failure_data:
                    empty_df = pd.DataFrame(columns=['提示'])
                    empty_df.loc[0] = ['无数据']
                    empty_df.to_excel(writer, index=False, sheet_name='数据')
            
            logger.info(f"✅ 结果已保存到Excel文件，成功数据在success工作表，失败数据在failure工作表：{output_path}")
            return True
        except Exception as e:
            logger.info(f"❌ 保存结果到Excel文件失败：{str(e)}")
            return False
    
    def _apply_excel_formatting(self, worksheet, header_font, header_fill, header_alignment, thin_border, normal_alignment):
        """应用Excel格式到工作表"""
        # 设置表头格式
        for cell in worksheet[1]:  # 第一行为表头
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置所有单元格的边框和对齐方式
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = normal_alignment
        
        # 调整列宽以适应内容
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            
            for cell in col:
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # 设置列宽，添加一些额外空间
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
            worksheet.column_dimensions[column].width = adjusted_width

    def collect_data_for_all_buildings(self) -> Dict[str, Dict[str, Dict[str, Any]]]:
        """为所有楼栋收集数据"""
        # 获取所有楼栋JSON文件
        building_files = []
        for file in os.listdir(self.resource_dir):
            if file.endswith('_data.json') and not file.endswith('_improved_data_collected_'):
                building_files.append(file)
        
        logger.info(f"🚀 开始为所有楼栋收集数据，共{len(building_files)}个楼栋")
        
        # 为每个楼栋收集数据
        all_results = {}
        for building_file in building_files:
            logger.info(f"\n🔄 处理楼栋：{building_file}")
            building_results = self.collect_data_for_building(building_file)
            if building_results:
                all_results[building_file] = building_results
                # collect_data_for_building方法内部已经调用了save_results_to_json
                # 此处不再重复调用
        
        logger.info(f"\n✅ 所有楼栋数据收集完成，共处理{len(all_results)}/{len(building_files)}个楼栋")
        
        return all_results

# 示例用法
if __name__ == "__main__":
    import argparse
    import glob
    
    # 创建参数解析器
    parser = argparse.ArgumentParser(description='数据收集管理器 - 收集PLC累计制热制冷量数据')
    parser.add_argument('-f', '--file', type=str, help='指定resource目录下的文件名，支持通配符，例如：1* 或 *data.json')
    
    # 解析命令行参数
    args = parser.parse_args()
    
    # 创建数据收集管理器，设置线程池大小
    manager = ImprovedDataCollectionManager(max_workers=10)
    manager.start()
    
    try:
        # 确定要处理的文件列表
        if args.file:
            # 使用通配符匹配文件
            pattern = os.path.join(manager.resource_dir, args.file)
            building_files = glob.glob(pattern)
            
            # 提取文件名部分（不包含路径）
            building_files = [os.path.basename(f) for f in building_files]
            
            # 过滤，只保留以_data.json结尾的文件
            building_files = [f for f in building_files if f.endswith('_data.json')]
            
            if not building_files:
                logger.info(f"❌ 未找到匹配的文件: {args.file}")
                # 显示帮助信息
                parser.print_help()
        else:
            # 如果没有指定文件，默认使用3#_data.json
            building_files = ['1#_data.json']
            logger.info("⚠️  未指定文件名，默认使用3#_data.json")
            logger.info("💡 使用 -f 参数指定文件名，例如: python improved_data_collection_manager.py -f 1* 或 python improved_data_collection_manager.py -f *data.json")
        
        # 处理每个匹配的文件
        logger.info(f"🚀 开始处理文件列表：{building_files}")
        for building_file in building_files:
            logger.info(f"\n🔄 处理文件：{building_file}")
            results = manager.collect_data_for_building(building_file)
            
            if results:
                logger.info("📋 收集到的数据概览:")
                success_count = sum(1 for device_data in results.values() if device_data['status'] == 'success')
                partial_count = sum(1 for device_data in results.values() if device_data['status'] == 'partial_success')
                failed_count = sum(1 for device_data in results.values() if device_data['status'] == 'failed')
                
                logger.info(f"  总设备数: {len(results)}")
                logger.info(f"  成功: {success_count}, 部分成功: {partial_count}, 失败: {failed_count}")
                
                # 只打印第一个设备的详细信息作为示例
                first_device = next(iter(results.items()), None)
                if first_device:
                    device_id, device_data = first_device
                    logger.info(f"  示例设备ID: {device_id}")
                    logger.info(f"  基本信息: {device_data['专有部分坐落']}, IP: {device_data['IP地址']}")
                    logger.info(f"  PLC IP: {device_data.get('PLC IP地址', 'N/A')}")
                    logger.info(f"  收集状态: {device_data['status']}")
                    logger.info(f"  数据参数数量: {len(device_data['data'])}")
                    logger.info("  ----------")
        
    except KeyboardInterrupt:
        logger.info("\n✅ 用户手动终止程序")
    except Exception as e:
        logger.info(f"\n❌ 程序异常：{str(e)}")
        # 打印完整的错误堆栈
        import traceback
        logger.info(traceback.format_exc())
    finally:
        # 确保停止线程池
        manager.stop()